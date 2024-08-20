import os
import pandas as pd
from datetime import datetime
import shutil
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import time
import openpyxl
import configparser
import signal

# Function to check if a file is accessible
def is_file_accessible(file_path, mode='r'):
    try:
        with open(file_path, mode):
            pass
    except IOError:
        return False
    return True

# Function to read configuration from config.ini
def read_config(config_file='config.ini'):
    config = configparser.ConfigParser()
    if not os.path.exists(config_file):
        raise FileNotFoundError(f"Configuration file {config_file} not found.")

    config.read(config_file)
    client_folder = config.get('Folders', 'ClientFolder')
    dev_folder = config.get('Folders', 'DevFolder')
    log_file = config.get('Files', 'LogFile')
    snapshot_file = config.get('Files', 'SnapshotFile')
    ignored_files = config.get('Ignored', 'IgnoredFiles').split(',')
    return client_folder, dev_folder, log_file, snapshot_file, ignored_files

# Function to check and create necessary folders
def check_and_create_folders(client_folder, dev_folder):
    os.makedirs(client_folder, exist_ok=True)
    os.makedirs(dev_folder, exist_ok=True)
    print(f"Ensured that folders exist:\nClient: {client_folder}\nDev: {dev_folder}")

# Function to create the initial Excel file with required sheets if they do not exist
def create_initial_excel(file_path='snapshot.xlsx'):
    if not os.path.exists(file_path):
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            input_df = pd.DataFrame({
                'Client Folder': ['./ClientFolder'],
                'Dev Folder': ['./DevFolder']
            })
            input_df.to_excel(writer, sheet_name='Input', index=False)

            ignored_files_df = pd.DataFrame({
                'Ignored Files': ['ignore.txt']
            })
            ignored_files_df.to_excel(writer, sheet_name='Ignored Files', index=False)

            log_df = pd.DataFrame(columns=['Timestamp', 'Event'])
            log_df.to_excel(writer, sheet_name='Log', index=False)

        print(f"Initial Excel file created at {file_path}")

# Function to ensure that the Excel file has the required sheets
def ensure_excel_sheets(file_path='snapshot.xlsx'):
    if not is_file_accessible(file_path, 'a'):
        raise PermissionError(f"File {file_path} is not accessible. Please close it if it is open in another program.")

    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        workbook = writer.book
        if 'Input' not in workbook.sheetnames:
            input_df = pd.DataFrame({
                'Client Folder': ['./ClientFolder'],
                'Dev Folder': ['./DevFolder']
            })
            input_df.to_excel(writer, sheet_name='Input', index=False)
            print(f"Created 'Input' sheet in {file_path}")

        if 'Ignored Files' not in workbook.sheetnames:
            ignored_files_df = pd.DataFrame({
                'Ignored Files': ['ignore.txt']
            })
            ignored_files_df.to_excel(writer, sheet_name='Ignored Files', index=False)
            print(f"Created 'Ignored Files' sheet in {file_path}")

        if 'Log' not in workbook.sheetnames:
            log_df = pd.DataFrame(columns=['Timestamp', 'Event'])
            log_df.to_excel(writer, sheet_name='Log', index=False)
            print(f"Created 'Log' sheet in {file_path}")

# Function to read folder paths from the Excel file
def read_folder_paths(file_path='snapshot.xlsx'):
    df = pd.read_excel(file_path, sheet_name='Input')
    return df['Client Folder'][0], df['Dev Folder'][0]

# Function to read the list of ignored files from the Excel file
def read_ignored_files(file_path='snapshot.xlsx'):
    try:
        df = pd.read_excel(file_path, sheet_name='Ignored Files')
        return df['Ignored Files'].tolist()
    except Exception as e:
        print(f"Error reading ignored files: {e}")
        return []

# Function to create a snapshot of files in a folder
def create_snapshot(folder_path, ignored_files):
    snapshot = []
    for file_name in os.listdir(folder_path):
        if file_name in ignored_files:
            continue
        file_path = os.path.join(folder_path, file_name)
        if os.path.isfile(file_path):
            modified_time = os.path.getmtime(file_path)
            snapshot.append((file_name, datetime.fromtimestamp(modified_time)))
    return snapshot

# Function to compare the snapshots of the two folders
def compare_snapshots(client_snapshot, dev_snapshot):
    client_df = pd.DataFrame(client_snapshot, columns=['File Name', 'Client Modified'])
    dev_df = pd.DataFrame(dev_snapshot, columns=['File Name', 'Dev Modified'])

    combined_df = pd.merge(client_df, dev_df, on='File Name', how='outer', indicator=True)
    combined_df['Status'] = combined_df['_merge'].map({
        'both': 'Same File',
        'left_only': 'Modified in Client Only',
        'right_only': 'Modified in Dev Only'
    })

    combined_df.sort_values(by=['Client Modified', 'Dev Modified'], ascending=False, inplace=True)
    return combined_df

# Function to synchronize files based on the comparison
def sync_files(combined_df, client_folder, dev_folder):
    for _, row in combined_df.iterrows():
        file_name = row['File Name']
        client_file_path = os.path.join(client_folder, file_name)
        dev_file_path = os.path.join(dev_folder, file_name)

        if row['Status'] == 'Modified in Client Only':
            if os.path.exists(client_file_path):
                shutil.copy(client_file_path, dev_folder)
                print(f"Copied {file_name} from Client to Dev.")
            else:
                if os.path.exists(dev_file_path):
                    os.remove(dev_file_path)
                    print(f"Deleted {file_name} from Dev because it was deleted in Client.")
        elif row['Status'] == 'Same File':
            if os.path.exists(client_file_path) and not os.path.exists(dev_file_path):
                shutil.copy(client_file_path, dev_folder)
                print(f"Copied {file_name} from Client to Dev because it was missing in Dev.")

# Function to update the Excel file with the snapshot data and highlight differences
def update_snapshot_excel(combined_df, excel_path):
    temp_file = "temp_snapshot.xlsx"

    with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
        combined_df.to_excel(writer, sheet_name='Last Snapshot', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Last Snapshot']
        apply_conditional_formatting(combined_df, worksheet)

    if is_file_accessible(excel_path, 'a'):
        os.replace(temp_file, excel_path)  # Replace the original file with the updated one
    else:
        print(f"{excel_path} is currently open. Please close it if possible.")
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            combined_df.to_excel(writer, sheet_name='Last Snapshot', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Last Snapshot']
            apply_conditional_formatting(combined_df, worksheet)

    print(f"Snapshot updated in {excel_path}")

# Function to apply conditional formatting
def apply_conditional_formatting(combined_df, worksheet):
    for index, row in combined_df.iterrows():
        cell = worksheet.cell(row=index + 2, column=1)
        if row['Status'] == 'Same File':
            cell.fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00",
                                                    fill_type="solid")  # Green
        elif row['Status'] == 'Modified in Client Only':
            cell.fill = openpyxl.styles.PatternFill(start_color="FFCC00", end_color="FFCC00",
                                                    fill_type="solid")  # Yellow
        elif row['Status'] == 'Modified in Dev Only':
            cell.fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red

# Function to log changes to the text file
def log_changes(log_file, event):
    if not is_file_accessible(log_file, 'a'):
        print(f"{log_file} is not accessible. Please check permissions.")
        return

    with open(log_file, 'a') as f:
        f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {event}\n")

    print(f"Logged event: {event}")

# Event handler class for monitoring folder changes
class FileChangeHandler(FileSystemEventHandler):
    def __init__(self, client_folder, dev_folder, snapshot_file, ignored_files, log_file):
        self.client_folder = client_folder
        self.dev_folder = dev_folder
        self.snapshot_file = snapshot_file
        self.ignored_files = ignored_files
        self.log_file = log_file

    def on_any_event(self, event):
        print(f"Detected change: {event.event_type} - {event.src_path}")
        self.sync_and_log()

    def sync_and_log(self):
        client_snapshot = create_snapshot(self.client_folder, self.ignored_files)
        dev_snapshot = create_snapshot(self.dev_folder, self.ignored_files)
        combined_df = compare_snapshots(client_snapshot, dev_snapshot)
        sync_files(combined_df, self.client_folder, self.dev_folder)
        update_snapshot_excel(combined_df, self.snapshot_file)
        log_changes(self.log_file, f"Folders synced: {datetime.now()}")

# Function to handle graceful exit
def signal_handler(sig, frame):
    print("\nGraceful exit initiated...")
    observer.stop()
    observer.join()
    print("Exiting program...")
    exit(0)

# Main function to start the monitoring process
def main():
    config_file = 'config.ini'
    create_initial_excel()
    ensure_excel_sheets()

    client_folder, dev_folder, log_file, snapshot_file, ignored_files = read_config(config_file)
    check_and_create_folders(client_folder, dev_folder)

    event_handler = FileChangeHandler(client_folder, dev_folder, snapshot_file, ignored_files, log_file)
    global observer  # Make observer global so that it can be accessed in the signal_handler
    observer = Observer()
    observer.schedule(event_handler, path=client_folder, recursive=True)
    observer.schedule(event_handler, path=dev_folder, recursive=True)

    # Set up signal handler for graceful exit
    signal.signal(signal.SIGINT, signal_handler)

    print("Starting monitoring...")
    observer.start()

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        pass  # The signal_handler will handle the cleanup and exit

if __name__ == "__main__":
    main()
